# -*- coding: utf-8 -*-
'''=====================================
    @Author : Daniel
    @Date : 2022/3/25 10:40
    @Project : python_study
    @File : excel_reader.py
====================================='''
import xlrd,time,openpyxl

from openpyxl.styles import Alignment,PatternFill,colors,Border,Side
from 华测.API.httpFrame.model.http_model import http_model
from 华测.API.httpFrame.cfg.cfg import *

class readExcel:

    def __init__(self):
        self.wb = openpyxl.load_workbook(filepath)

    def my_border(self,t_border, b_border, l_border, r_border):
        '''定义边框样式'''
        border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                        bottom=Side(border_style=b_border, color=colors.BLACK),
                        left=Side(border_style=l_border, color=colors.BLACK),
                        right=Side(border_style=r_border, color=colors.BLACK))
        return border

    # def read(self):
    #     '''通过xlrd读取，原始代码'''
    #     models = []
    #     reader = xlrd.open_workbook(filepath)
    #     names = reader.sheet_names()
    #
    #     for name in names:
    #         case_xlsx = reader.sheet_by_name(name)
    #         for i in range(case_xlsx.nrows):
    #             if i == 0: #跳出第一个行
    #                 continue
    #             smart_lsit = []
    #             for j in range(case_xlsx.ncols):
    #                 smart_lsit.append(case_xlsx.cell(i,j).value)
    #             print(smart_lsit)
    #             model = http_model()
    #             # getattr(model,info)=smart_lsit[info]
    #             model.url = smart_lsit[0]
    #             model.method = smart_lsit[2]
    #             model.data = smart_lsit[4]
    #             model.headers = smart_lsit[3]
    #             model.extract = smart_lsit[8]
    #             model.assert_data = smart_lsit[5]
    #             model.assert_options = smart_lsit[6]
    #             model.assert_value = smart_lsit[7]
    #             model.is_need = smart_lsit[9]
    #             model.last_value = smart_lsit[10]
    #             model.req_params_type = smart_lsit[11]
    #             model.desc = smart_lsit[1]
    #
    #             models.append(model)
    #     print(models)
    #     return models
    #
    # def read_data(self):
    #     '''通过xlrd读取，修改后的代码'''
    #     models = []
    #     reader = xlrd.open_workbook(filepath)
    #     names = reader.sheet_names()
    #     for name in names:
    #         case_xlsx = reader.sheet_by_name(name)
    #         for i in range(case_xlsx.nrows):
    #             smart_lsit = []
    #             data_dic = {}
    #             excel_num = {} # 存储结果、测试时间、测试人员的表格号-->   {result_num:[1,2],test_date_num:[2,2]}
    #             if i == 0: #跳出第一个行
    #                 continue
    #             for j in range(case_xlsx.ncols):
    #                 for value in ['result','test_date','tester']:
    #                     if case_xlsx.cell(0,j).value == value:
    #                         excel_num.update({'{}_num'.format(value): (i+1, j+1)}) # openpyxl 序号不同，所以需要+1
    #                 if excel_num != {}:
    #                     data_dic.update({'excel_num':excel_num})
    #
    #                 data_dic.update({case_xlsx.cell(0,j).value:case_xlsx.cell(i,j).value})
    #             print(data_dic)
    #             smart_lsit.append(data_dic)
    #
    #             model = http_model()
    #             for j in list(i for i in model.__dir__() if i[:2] !='__'): #遍历model自定义属性列表
    #                 if j == "url":
    #                     model.url = url+smart_lsit[0]['api']
    #                 else:
    #                     setattr(model,j,smart_lsit[0][j])
    #             models.append(model)
    #     print(models[0].excel_num)
    #     return models

    def read_data_op(self):
        '''通过openpyxl读取'''
        models = []
        names = self.wb.sheetnames
        for name in names:
            ws = self.wb[name]
            for i in range(2,ws.max_row+1): # 过滤第一行
                smart_lsit = []
                data_dic = {}
                excel_num = {}  # 存储结果、测试时间、测试人员的表格号-->   {result_num:[1,2],test_date_num:[2,2]}
                for j in range(1,ws.max_column+1):
                    for value in ['result', 'test_date', 'tester']:
                        if ws.cell(1, j).value == value:
                            excel_num.update({'{}_num'.format(value): (i, j)})
                    if excel_num != {}:
                        data_dic.update({'excel_num': excel_num})

                    data_dic.update({ws.cell(1, j).value: ws.cell(i, j).value})
                # print(data_dic)
                smart_lsit.append(data_dic)

                model = http_model()
                for j in list(i for i in model.__dir__() if i[:2] != '__'):  # 遍历model自定义属性列表
                    if j == "url":
                        model.url = url + smart_lsit[0]['api']
                    else:
                        setattr(model, j, smart_lsit[0][j])
                models.append(model)
        # print(models[0].excel_num)
        return models

    def write_value(self, row, col, value):
        ws = self.wb[self.wb.sheetnames[0]]
        ws.cell(row, col).value = value
        ws.cell(row,col).alignment=Alignment(horizontal='center',vertical='center')
        if value.upper() == 'PASS':
            ws.cell(row,col).fill=PatternFill("solid", fgColor='3CB371') #设置通过的背景色为绿色，fgColor需使用16进制RGB颜色
        elif value.upper() == 'FAILD':
            ws.cell(row, col).fill = PatternFill("solid", fgColor='CD3333') # 失败的背景色为红色

        ws.cell(row, col).border=self.my_border('thin', 'thin', 'thin', 'thin') # 设置边框为所有框线

        self.wb.save(savepath+r'\testResult-{}.xlsx'.format(self.get_date()))

    def get_date(self):
        return time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))

if __name__ == '__main__':
    r = readExcel()
    # r.read()
    # r.read_data()
    # r.read_data_op()
    r.write_value(2,15,'PasS')
    r.write_value(2,16,'FAILD')
    r.write_value(3,16,'aaaa')
    r.write_value(4,16,'nnnn')
