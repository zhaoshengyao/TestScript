# -*- coding: utf-8 -*-
'''=====================================
    @Author : Daniel
    @Date : 2021/9/26 10:54
    @Project : Project
    @File : readExcel.py
====================================='''
from openpyxl import *

from collections import Counter
from openpyxl.utils import column_index_from_string


class excelOp():
    def __init__(self, filename=None,sheetname=None):
        self.file = filename
        self.wb = load_workbook(self.file)
        # sheets = self.wb.sheetnames
        # self.sheet = sheets[0]
        self.ws = self.wb[sheetname]

    def get_row_clo_num(self):
        '''获取表格的总行数和总列数'''
        totalRows = self.ws.max_row
        totalColumns = self.ws.max_column
        return totalRows, totalColumns

    def get_cell_value(self, row, column):
        '''获取某个单元格的值'''
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    def get_col_value(self, column):
        '''获取某列的所有值'''
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data


    def get_row_value(self, row):
        '''获取某行所有值'''
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data


    def set_cell_value(self, row, colunm, cellvalue):
        '''设置某个单元格的值'''
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    def count_value(self,column,value):
        '''
        统计某列中某元素出现的次数
        :param column:要统计的列
        :param value:要统计的值
        :return:
        '''
        value_list = []
        for i in self.get_col_value(column):
            if value in str(i) :
                value_list.append(i)

        return Counter(value_list)[value]

    def search_value(self,keyword,min_col=2,max_col=2):
        '''
        获取元素的坐标
        :param keyword: 输入字符串
        :param min_col: 最小列
        :param max_col: 最大列
        :return:
        '''
        col_list = []
        for row in self.ws.iter_rows(min_col=min_col,max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    info = cell.value.find(keyword)
                    if info == 0:
                        # print(cell.value) #获取keyword的列号
                        # print(cell.row) #获取keyword的行号
                        # print(cell.column) #获取keyword的列号
                        col_list.append(cell.row)

        return col_list