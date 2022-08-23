# -*- coding: utf-8 -*-
'''=====================================
    @Author : Daniel
    @Date : 2022/3/25 10:40
    @Project : python_study
    @File : excel_reader.py
====================================='''
import time, openpyxl

from openpyxl.styles import Alignment, PatternFill, colors, Border, Side
from DiaoApp.model.http_model import http_model
from DiaoApp.cfg.cfg import *

os.chdir(os.getcwd())


class readExcel:

    def __init__(self):
        self.wb = openpyxl.load_workbook(config['filepath'])

    def my_border(self, t_border, b_border, l_border, r_border):
        '''定义边框样式'''
        border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                        bottom=Side(border_style=b_border, color=colors.BLACK),
                        left=Side(border_style=l_border, color=colors.BLACK),
                        right=Side(border_style=r_border, color=colors.BLACK))
        return border

    def read_data_op(self):
        '''通过openpyxl读取'''
        models = []
        names = self.wb.sheetnames
        for name in names:
            ws = self.wb[name]
            big_smart_list = []
            try:
                for i in range(3, ws.max_row + 1):  # 过滤第一行
                    smart_list = []  # 用例数据
                    data_dic = {}  # 用例字典
                    tester_msg = {}  # 存储结果、测试时间、测试人员的表格号-->   {result_num:[1,2],test_date_num:[2,2]}

                    for j in range(1, ws.max_column + 1):
                        for value in ['result', 'test_date', 'tester']:
                            if ws.cell(2, j).value == value:
                                tester_msg.update({'{}_num'.format(value): (i, j)})

                        if tester_msg != {}:
                            data_dic.update({'tester_msg': tester_msg})
                            data_dic.update({'sheet_name': name})
                        data_dic.update({ws.cell(2, j).value: ws.cell(i, j).value})
                    if len([i for i in list(data_dic.values())[:3] if i is None]) < 3:  # 过滤Excel前三列为空的无效用例
                        if data_dic['is_run'] is True:  # 只添加需要执行的数据
                            smart_list.append(data_dic)
                            big_smart_list.append(data_dic)

                            # 判断是否存在依赖用例
                            if data_dic['dependence_case'] is not None:
                                for case_list in big_smart_list:
                                    for dependence in data_dic['dependence_case'].split(","):
                                        if data_dic['dependence_case'] == case_list['case_num']:  # 依赖用例编号==用例编号
                                            smart_list.insert(0, case_list)

                            # 实例化model，生成用例对象
                            model = http_model()

                            for data in smart_list:
                                for j in list(i for i in model.__dir__() if i[:2] != '__'):  # 遍历model自定义属性列表
                                    if j == "url":
                                        model.url = config['url'] + '/' + data['api']

                                    else:
                                        setattr(model, j, data[j])
                            models.append(model)

            except Exception as e:
                print("读取数据失败，请检查！", e)
        return models

    def write_value(self, sheetName, row, col, value):
        if not os.path.exists(config['savepath']):
            os.mkdir(config['savepath'])
        ws = self.wb[sheetName]
        ws.cell(row, col).value = value
        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        if value.upper() == 'PASS':
            ws.cell(row, col).fill = PatternFill("solid", fgColor='3CB371')  # 设置通过的背景色为绿色，fgColor需使用16进制RGB颜色
        elif value.upper() == 'FAILD':
            ws.cell(row, col).fill = PatternFill("solid", fgColor='CD3333')  # 失败的背景色为红色
        elif 'Error' in value:
            ws.cell(row, col).fill = PatternFill("solid", fgColor='FFFF77')

        ws.cell(row, col).border = self.my_border('thin', 'thin', 'thin', 'thin')  # 设置边框为所有框线

    def get_date(self):
        # return time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        return "DiAoAPPTestReport"

    def save(self):
        return self.wb.save(config['savepath'] + r'\testResult-{}.xlsx'.format(self.get_date()))

# if __name__ == '__main__':
#     r = readExcel()
#     r.read_data_op()
# print(len(r.read_data_op()))
